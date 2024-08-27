import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

#Variáveis para a tabela de gastps
tabela = pd.DataFrame(columns=["Categoria", "Descrição", "Valor", "Pago", "Data de Pagamento"])

alteracoes_salvas = True

#Função que eu sofri horrores para formatar valores em reais(ainda não da certo 100%)
def formatar_valor(valor):
    return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

# Função para remover formatação e converter string para float
def converter_para_float(valor_str):
    valor_str = valor_str.replace('R$ ', '').replace('.', '').replace(',', '.')
    try:
        return float(valor_str)
    except ValueError:
        print(f"Valor não pode ser convertido: {valor_str}")
        return 0.0

#Aqui eu add um gasto, botão
def adicionar_gasto(categoria, descricao, valor, pago, data_pagamento):
    global tabela, alteracoes_salvas
    try:
        valor_float = converter_para_float(valor)
        valor_formatado = formatar_valor(valor_float)
        novo_gasto = pd.DataFrame({
            "Categoria": [categoria],
            "Descrição": [descricao],
            "Valor": [valor_formatado],
            "Pago": [pago],
            "Data de Pagamento": [data_pagamento]
        })
        tabela = pd.concat([tabela, novo_gasto], ignore_index=True)
        atualizar_tabela()
        atualizar_total()
        alteracoes_salvas = False
    except ValueError:
        print("O valor inserido não é válido. Certifique-se de que é um número.")

#Aqui eu carrego gastos pra edição, botão
def carregar_gasto_para_edicao():
    selected_item = tree.focus()
    if selected_item:
        valores = tree.item(selected_item, 'values')
        entry_categoria.delete(0, tk.END)
        entry_categoria.insert(0, valores[0])
        entry_descricao.delete(0, tk.END)
        entry_descricao.insert(0, valores[1])
        entry_valor.delete(0, tk.END)
        entry_valor.insert(0, valores[2])
        pago_var.set(valores[3] == '✔')
        entry_data_pagamento.delete(0, tk.END)
        entry_data_pagamento.insert(0, valores[4])

#Aqui eu edito um gasto existente, botão
def editar_gasto_interface():
    global alteracoes_salvas
    selected_item = tree.focus()
    if selected_item:
        tabela.loc[int(selected_item), 'Categoria'] = entry_categoria.get()
        tabela.loc[int(selected_item), 'Descrição'] = entry_descricao.get()
        tabela.loc[int(selected_item), 'Valor'] = formatar_valor(converter_para_float(entry_valor.get()))
        tabela.loc[int(selected_item), 'Pago'] = pago_var.get()
        tabela.loc[int(selected_item), 'Data de Pagamento'] = entry_data_pagamento.get()
        atualizar_tabela()
        atualizar_total()
        alteracoes_salvas = False

#Aqui eu consigo excluir algum gasto criado, botão
def excluir_gasto_interface():
    global alteracoes_salvas
    selected_item = tree.focus()
    if selected_item:
        tabela.drop(int(selected_item), inplace=True)
        tabela.reset_index(drop=True, inplace=True)
        atualizar_tabela()
        atualizar_total()
        alteracoes_salvas = False

#Aqui eu salvo a planilha no excel, botão
def salvar_planilha(mes_ano):
    global alteracoes_salvas
    try:
        tabela.to_excel(f"gastos_{mes_ano}.xlsx", index=False)
        personalizar_planilha(f"gastos_{mes_ano}.xlsx")
        print(f"Planilha do mês {mes_ano} salva com sucesso.")
        alteracoes_salvas = True
    except Exception as e:
        print(f"Erro ao salvar a planilha: {e}")

#Consigo pesoanlizar a planilha que já foi salva
def personalizar_planilha(arquivo):
    try:
        wb = load_workbook(arquivo)
        ws = wb.active

        #Só enfeite, alterando a fonte e a cor de fundo do cabeçalho
        for row in ws.iter_rows(min_row=1, max_row=1, max_col=5):
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="007ACC", end_color="007ACC", fill_type="solid")

        #Só enfeite, ajustando a largura das colunas
        ws.column_dimensions['A'].width = 30  # Categoria
        ws.column_dimensions['B'].width = 40  # Descrição
        ws.column_dimensions['C'].width = 20  # Valor
        ws.column_dimensions['D'].width = 10  # Pago
        ws.column_dimensions['E'].width = 20  # Data de Pagamento

        #Só enfeite, add cor de fundo vermelho claro para a coluna "Valor"
        fill_valor = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
            row[2].fill = fill_valor  # Índice 2 é a coluna "Valor"

        #Tõ substituindo os valores booleanos pra checkmark, fica mais bonitinho
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5):
            for cell in row:
                if cell.column == 4:  # Coluna "Pago"
                    if cell.value == True:
                        cell.value = '✔'
                    elif cell.value == False:
                        cell.value = ''

        wb.save(arquivo)
    except Exception as e:
        print(f"Erro ao personalizar a planilha: {e}")

#Carrega dados da planilha de um mês anterior, botão
def carregar_planilha_anterior():
    global tabela, alteracoes_salvas
    arquivo = filedialog.askopenfilename(title="Selecionar Planilha do Mês Anterior",
                                         filetypes=[("Arquivos Excel", "*.xlsx")])
    if arquivo:
        try:
            tabela = pd.read_excel(arquivo)

            #Aqui eu to tratando a coluna valor, ainda to sofrendo muito aqui(não funciona 100%)
            def converter_valor(valor):
                try:
                    valor = str(valor).replace('R$', '').replace('.', '').replace(',', '.')
                    return float(valor) if valor else 0.0
                except ValueError:
                    return 0.0

            tabela['Valor'] = tabela['Valor'].apply(converter_valor)

            #Aqui to tratando a coluna pago, pois não tava aparecendo o checkmark
            tabela['Pago'] = tabela['Pago'].fillna(False).astype(bool)

            #Aqui to tratando a coluna data de pagamento, pois tava aparecendo o que não devia, agora aparece vazia
            tabela['Data de Pagamento'] = tabela['Data de Pagamento'].fillna('')

            atualizar_tabela()
            atualizar_total()
            print(f"Planilha carregada com sucesso de {arquivo}.")
            alteracoes_salvas = True
        except Exception as e:
            print(f"Erro ao carregar a planilha: {e}")
    else:
        print("Nenhum arquivo selecionado.")


#Atualiza a exibição da tabela na interface
def atualizar_tabela():
    for row in tree.get_children():
        tree.delete(row)
    for index, row in tabela.iterrows():
        tree.insert("", tk.END, iid=index, values=list(row))

#Atualiza o total de gastos
def atualizar_total():
    total = tabela["Valor"].sum()
    label_total.config(text=f"Total de Gastos: {formatar_valor(total)}")

#Adiciona um gasto pelo botão da interface
def adicionar_gasto_interface():
    categoria = entry_categoria.get()
    descricao = entry_descricao.get()
    valor = entry_valor.get()
    pago = pago_var.get()
    data_pagamento = entry_data_pagamento.get()
    adicionar_gasto(categoria, descricao, valor, pago, data_pagamento)
    limpar_campos()

#Limpa os campos de entrada, para não ter que apagar manualmente
def limpar_campos():
    entry_categoria.delete(0, tk.END)
    entry_descricao.delete(0, tk.END)
    entry_valor.delete(0, tk.END)
    pago_var.set(False)
    entry_data_pagamento.delete(0, tk.END)

#COnfirma se quer fechar o programa caso tenha alterações não salvas
def confirmar_saida():
    if not alteracoes_salvas:
        if messagebox.askokcancel("Sair", "Você tem alterações não salvas. Tem certeza que deseja fechar sem salvar?"):
            root.destroy()
    else:
        root.destroy()

#Configura a interface
root = tk.Tk()
root.title("Controle de Gastos")

#Fechamento de janela só com confirmação de saída
root.protocol("WM_DELETE_WINDOW", confirmar_saida)

#Escrever Mês e Ano
tk.Label(root, text="Mês e Ano (ex: Ago2024):").grid(row=0, column=0)
entry_mes_ano = tk.Entry(root)
entry_mes_ano.grid(row=0, column=1)
tk.Button(root, text="Salvar Planilha", command=lambda: salvar_planilha(entry_mes_ano.get())).grid(row=0, column=2)
tk.Button(root, text="Carregar Planilha do Mês Anterior", command=carregar_planilha_anterior).grid(row=0, column=3)

#Entradas para adicionar um novo gasto
tk.Label(root, text="Categoria").grid(row=1, column=0)
entry_categoria = tk.Entry(root)
entry_categoria.grid(row=1, column=1)

tk.Label(root, text="Descrição").grid(row=2, column=0)
entry_descricao = tk.Entry(root)
entry_descricao.grid(row=2, column=1)

tk.Label(root, text="Valor").grid(row=3, column=0)
entry_valor = tk.Entry(root)
entry_valor.grid(row=3, column=1)

tk.Label(root, text="Pago").grid(row=4, column=0)
pago_var = tk.BooleanVar()
check_pago = tk.Checkbutton(root, variable=pago_var)
check_pago.grid(row=4, column=1)

tk.Label(root, text="Data de Pagamento").grid(row=5, column=0)
entry_data_pagamento = tk.Entry(root)
entry_data_pagamento.grid(row=5, column=1)

#Os botões que adicionam, editam e excluem os gastos
tk.Button(root, text="Adicionar Gasto", command=adicionar_gasto_interface).grid(row=6, column=0)
tk.Button(root, text="Editar Gasto", command=editar_gasto_interface).grid(row=6, column=1)
tk.Button(root, text="Excluir Gasto", command=excluir_gasto_interface).grid(row=6, column=2)
tk.Button(root, text="Carregar Gasto para Edição", command=carregar_gasto_para_edicao).grid(row=6, column=3)

#Tabela de rolagem pra rolar bonitinho e poder percorrer pelos gastos
frame_tree = tk.Frame(root)
frame_tree.grid(row=7, column=0, columnspan=4, pady=10, sticky="nsew")

tree_scroll = tk.Scrollbar(frame_tree)
tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

tree = ttk.Treeview(frame_tree, columns=('Categoria', 'Descrição', 'Valor', 'Pago', 'Data de Pagamento'), show='headings', yscrollcommand=tree_scroll.set)
tree.heading('Categoria', text='Categoria')
tree.heading('Descrição', text='Descrição')
tree.heading('Valor', text='Valor')
tree.heading('Pago', text='Pago')
tree.heading('Data de Pagamento', text='Data de Pagamento')

#Só enfeite pra ajustar a largura das colunas
tree.column('Categoria', width=150)
tree.column('Descrição', width=200)
tree.column('Valor', width=150)
tree.column('Pago', width=80)
tree.column('Data de Pagamento', width=150)

tree.pack(fill=tk.BOTH, expand=True)
tree_scroll.config(command=tree.yview)

#Mostra o total de gastos no rodapé
label_total = tk.Label(root, text="Total de Gastos: R$ 0.00")
label_total.grid(row=8, column=0, columnspan=4, pady=10)

#Inicia o programa
root.mainloop()
